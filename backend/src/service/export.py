import csv
import io
import logging
from collections.abc import AsyncGenerator, Sequence
from typing import Any, TypedDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from core.model import EventRecord
from schema.records import RecordData

logger = logging.getLogger(__name__)

COLUMN_MAPPING: dict[str, str] = {
    "country": "Country",
    "region": "Region",
    "district": "District",
    "locality": "Locality",
    "is_manual_location": "Manual Location",
    "latitude": "Latitude",
    "longitude": "Longitude",
    "verbatimcoordinates": "Verbatim Coordinates",
    "coordinate_uncertainty": "Coordinate Uncertainty (m)",
    "georeferencedby": "Geo Referenced By",
    "location_remarks": "Location Remarks",
    "verbatim_date": "Date",
    "date_precision": "Date Precision",
    "is_interval": "Date Interval",
    "habitat": "Habitat",
    "sampling_protocol": "Sampling Protocol",
    "sampling_effort": "Sampling Effort",
    "sample_size_value": "Sample Size Value",
    "sample_size_unit": "Sample Size Unit",
    "event_remarks": "Event Remarks",
    "field_number": "Field Number",
    "catalog_number": "Catalog Number",
    "collection_code": "Collection Code",
    "recorded_by": "Recorded By",
    "family": "Family",
    "genus": "Genus",
    "species": "Species",
    "tax_verbatim": "Taxon Verbatim",
    "taxon_rank": "Taxon Rank",
    "type_status": "Type Status",
    "accepted_name": "Accepted Name",
    "taxon_remarks": "Taxon Remarks",
    "quantity": "Quantity",
    "quantity_type": "Quantity Type",
    "sex": "Sex",
    "life_stage": "Life Stage",
    "occurrence_remarks": "Occurrence Remarks",
    "identification_remarks": "Identification Remarks",
}


REVERSE_COLUMN_MAPPING: dict[str, str] = {v: k for k, v in COLUMN_MAPPING.items()}


class ParseResult(TypedDict):
    success: bool
    record: RecordData | None
    error: ValidationError | None


def is_row_empty(row: dict[str, Any | None]) -> bool:
    return all(v is None or str(v).strip() == "" for v in row.values())


def _row_to_record_data(row: dict[str, str | None]) -> ParseResult:
    """Convert a row dict to RecordData using pydantic validation."""
    data_dict: dict[str, str] = {}
    for display_name, field in REVERSE_COLUMN_MAPPING.items():
        raw_value = row.get(display_name)
        if raw_value is not None:
            data_dict[field] = raw_value
    try:
        record = RecordData.model_validate(data_dict)
        return {"success": True, "record": record, "error": None}
    except ValidationError as e:
        return {"success": False, "record": None, "error": e}


async def read_excel(file_content: bytes) -> AsyncGenerator[ParseResult]:
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
    ws = wb.active
    if ws is None:
        return
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell) if cell is not None else "" for cell in row]
        else:
            row_dict: dict[str, str | None] = {
                headers[j]: (str(row[j]) if row[j] is not None else None)
                for j in range(len(headers))
                if j < len(row)
            }
            yield _row_to_record_data(row_dict)
    wb.close()


async def read_csv(file_content: bytes) -> AsyncGenerator[ParseResult, None]:
    text = file_content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        row_dict = {k: (v if v else None) for k, v in row.items()}
        yield _row_to_record_data(row_dict)


def records_to_excel(records: Sequence[EventRecord]) -> bytes:
    wb = Workbook()
    ws = wb.active

    if ws is None:
        logger.error("ws is None")
        raise Exception

    headers = [COLUMN_MAPPING[field] for field in COLUMN_MAPPING]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.cell(row=1, column=col).font = Font(bold=True)

    for record in records:
        row = [getattr(record, field, None) for field in COLUMN_MAPPING]
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def records_to_csv(records: Sequence[EventRecord]) -> str:
    output = io.StringIO()
    fieldnames = [COLUMN_MAPPING[field] for field in COLUMN_MAPPING]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for record in records:
        row = {
            COLUMN_MAPPING[field]: getattr(record, field, None)
            for field in COLUMN_MAPPING
        }
        writer.writerow(row)

    return output.getvalue()
