import csv
import io
import logging
from typing import Annotated, TypedDict
from uuid import uuid4

from fastapi import APIRouter, Depends, File, UploadFile
from openpyxl import load_workbook

from core.config import settings
from core.dependencies import DBSession, TokenUser
from core.exceptions import RecordLimitExceededError
from core.model import EventRecord
from repository.user import get_user_expect
from schema.records import RecordData
from service.export import COLUMN_MAPPING
from service.records import RecordService, _mock_validate_record

router = APIRouter(
    prefix="/records",
)

logger = logging.getLogger(__name__)


class ImportResult(TypedDict):
    imported: int
    failed: int


REVERSE_COLUMN_MAPPING: dict[str, str] = {v: k for k, v in COLUMN_MAPPING.items()}

BOOLEAN_TRUE_VALUES: frozenset[str] = frozenset({"TRUE", "YES", "1", "T", "Y"})
BOOLEAN_FALSE_VALUES: frozenset[str] = frozenset({"FALSE", "NO", "0", "F", "N"})

FLOAT_FIELDS: frozenset[str] = frozenset({
    "latitude",
    "longitude",
    "coordinate_uncertainty",
    "sample_size_value",
    "quantity",
})

BOOLEAN_FIELDS: frozenset[str] = frozenset({
    "is_manual_location",
    "is_interval",
    "tax_verbatim",
})


def _convert_value(field: str, value: object) -> object:
    if value is None or value == "":
        return None

    str_value = str(value).strip()

    if field in BOOLEAN_FIELDS:
        return _convert_bool(str_value)

    if field in FLOAT_FIELDS:
        try:
            return float(str_value)
        except ValueError:
            return None

    return str_value


def _convert_bool(str_value: str) -> bool | None:
    upper = str_value.upper()
    if upper in BOOLEAN_TRUE_VALUES:
        return True
    if upper in BOOLEAN_FALSE_VALUES:
        return False
    return None


def _row_to_record_data(row: dict[str, str | None], publ_id: int) -> RecordData:
    data = RecordData(publ_id=publ_id)
    for display_name, field in REVERSE_COLUMN_MAPPING.items():
        raw_value = row.get(display_name)
        converted = _convert_value(field, raw_value)
        if converted is not None:
            setattr(data, field, converted)
    return data


def _is_row_empty(row: dict[str, str | None]) -> bool:
    return all(v is None or str(v).strip() == "" for v in row.values())


async def _read_excel(file_content: bytes) -> list[dict[str, str | None]]:
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows: list[dict[str, str | None]] = []
    headers: list[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell) if cell is not None else "" for cell in row]
        else:
            row_dict: dict[str, str | None] = {
                headers[j]: (str(row[j]) if row[j] is not None else None)
                for j in range(len(headers))
                if j < len(row)
            }
            rows.append(row_dict)

    wb.close()
    return rows


async def _read_csv(file_content: bytes) -> list[dict[str, str | None]]:
    text = file_content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [{k: (v if v else None) for k, v in row.items()} for row in reader]


@router.post("/import")
async def import_records(
    file: Annotated[UploadFile, File(...)],
    token: TokenUser,
    session: DBSession,
    service: Annotated[RecordService, Depends()],
) -> ImportResult:
    try:
        content = await file.read()
    except Exception:
        return {"imported": 0, "failed": 0}

    if len(content) > settings.MAX_IMPORT_FILE_BYTES:
        return {"imported": 0, "failed": 0}

    filename = file.filename or ""
    if filename.endswith(".xlsx"):
        rows = await _read_excel(content)
    elif filename.endswith(".csv"):
        rows = await _read_csv(content)
    else:
        return {"imported": 0, "failed": 0}

    records_data: list[tuple[RecordData, str | None, str]] = []
    failed_count = 0

    user = await get_user_expect(session, token.user_id)
    if user.publ_id is None:
        return {"imported": 0, "failed": 0}

    publ_id = user.publ_id

    for row in rows:
        if _is_row_empty(row):
            continue

        record_data = _row_to_record_data(row, publ_id)
        errors = _mock_validate_record(record_data)
        record_type = "rec_ok" if errors is None else "rec_fail"

        records_data.append((record_data, errors, record_type))
        if errors is not None:
            failed_count += 1

    try:
        await service.check_import_limit(publ_id, len(records_data))
    except RecordLimitExceededError:
        return {"imported": 0, "failed": 0}

    event_records = [
        EventRecord(
            id=uuid4(),
            publ_id=publ_id,
            user_id=token.user_id,
            errors=errors,
            type=record_type,
            **r.model_dump(exclude_none=True, exclude={"publ_id"}),
        )
        for r, errors, record_type in records_data
    ]

    session.add_all(event_records)
    await session.commit()

    return {"imported": len(records_data) - failed_count, "failed": failed_count}
