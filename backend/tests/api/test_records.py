import csv
import io
from collections.abc import Callable
from datetime import UTC, datetime
from uuid import uuid4

import pytest
from conftest import SeedData
from fastapi import status
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from core.model import EventRecord


@pytest.mark.asyncio
async def test_create_record(
    authenticated_client: AsyncClient, seed_data: SeedData
) -> None:
    user = seed_data["users"][0]

    response = await authenticated_client.post(
        "/api/records",
        json={"publ_id": user.publ_id},
    )
    assert response.status_code == 201
    assert "id" in response.json()


@pytest.mark.asyncio
async def test_get_record(
    authenticated_client: AsyncClient, seed_data: SeedData
) -> None:
    user = seed_data["users"][0]
    record_id = seed_data["record_ids"][0]
    response = await authenticated_client.get(
        f"/api/records/{record_id}?user_id={user.user_id}"
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_records(
    authenticated_client: AsyncClient, seed_data: SeedData
) -> None:
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records?user_id={user.user_id}&publ_id={user.publ_id}&page=1&page_size=20"
    )
    assert response.status_code == 200
    data = response.json()
    assert "items" in data
    assert "total" in data
    assert "page" in data
    assert "page_size" in data
    assert isinstance(data["items"], list)


@pytest.mark.asyncio
async def test_update_record_set_coords(
    authenticated_client: AsyncClient, seed_data
) -> None:
    user = seed_data["users"][0]
    record_id = seed_data["record_ids"][2]
    response = await authenticated_client.put(
        f"/api/records/{record_id}?user_id={user.user_id}",
        json={"publ_id": user.publ_id, "latitude": 55.7, "longitude": 37.7},
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_update_record_clear_coords(
    authenticated_client: AsyncClient, seed_data
) -> None:
    user = seed_data["users"][0]
    record_id = seed_data["record_ids"][0]
    response = await authenticated_client.put(
        f"/api/records/{record_id}?user_id={user.user_id}",
        json={"publ_id": user.publ_id, "latitude": None, "longitude": None},
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_update_record_set_genus(
    authenticated_client: AsyncClient, seed_data
) -> None:
    user = seed_data["users"][0]
    record_id = seed_data["record_ids"][2]
    response = await authenticated_client.put(
        f"/api/records/{record_id}?user_id={user.user_id}",
        json={"publ_id": user.publ_id, "genus": "UpdatedGenus"},
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_update_record_clear_genus(
    authenticated_client: AsyncClient, seed_data
) -> None:
    user = seed_data["users"][0]
    record_id = seed_data["record_ids"][0]
    response = await authenticated_client.put(
        f"/api/records/{record_id}?user_id={user.user_id}",
        json={"publ_id": user.publ_id, "genus": None},
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_delete_record(
    authenticated_client: AsyncClient, seed_data: SeedData
) -> None:
    record_id = seed_data["record_ids"][0]
    response = await authenticated_client.delete(f"/api/records/{record_id}")
    assert response.status_code == status.HTTP_204_NO_CONTENT


@pytest.mark.asyncio
async def test_list_records_no_token(async_client: AsyncClient, seed_data: SeedData):
    user = seed_data["users"][0]
    response = await async_client.get(
        f"/api/records?user_id={user.user_id}&publ_id={user.publ_id}"
    )
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_get_record_wrong_user(
    authenticated_client_user2: AsyncClient, seed_data
):
    # user2 tries to access user 1's record
    user2 = seed_data["users"][1]
    record_id = seed_data["record_ids"][0]
    response = await authenticated_client_user2.get(
        f"/api/records/{record_id}?user_id={user2.user_id}"
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_records_all_publs(authenticated_client, seed_data: SeedData):
    user = seed_data["users"][0]
    response = await authenticated_client.get(f"/api/records?user_id={user.user_id}")
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_records_invalid_sort(authenticated_client, seed_data: SeedData):
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records?user_id={user.user_id}&publ_id={user.publ_id}&sort=invalid"
    )
    assert response.status_code == 422


@pytest.mark.asyncio
async def test_list_records_page_size_exceeds_max(
    authenticated_client, seed_data: SeedData
):
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records?user_id={user.user_id}&publ_id={user.publ_id}&page_size=200"
    )
    assert response.status_code == 422


@pytest.mark.asyncio
async def test_list_records_pagination_second_page(
    authenticated_client, seed_data: SeedData
):
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records?user_id={user.user_id}&publ_id={user.publ_id}&page=2&page_size=1"
    )
    assert response.status_code == 200
    data = response.json()
    assert data["page"] == 2


@pytest.mark.asyncio
async def test_list_records_sort_updated_at(authenticated_client, seed_data: SeedData):
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records?user_id={user.user_id}&publ_id={user.publ_id}&sort=updated_at"
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_get_record_not_found(authenticated_client, seed_data: SeedData):
    user = seed_data["users"][0]
    fake_uuid = "00000000-0000-0000-0000-000000000000"
    response = await authenticated_client.get(
        f"/api/records/{fake_uuid}?user_id={user.user_id}"
    )
    print(response)
    assert response.status_code == 404


@pytest.mark.asyncio
async def test_update_record_not_found(authenticated_client, seed_data: SeedData):
    user = seed_data["users"][0]
    fake_uuid = "00000000-0000-0000-0000-000000000000"
    response = await authenticated_client.put(
        f"/api/records/{fake_uuid}", json={"publ_id": user.publ_id, "latitude": 55.5}
    )
    assert response.status_code == 404


@pytest.mark.asyncio
async def test_get_record_invalid_uuid(authenticated_client, seed_data: SeedData):
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records/not-a-uuid?user_id={user.user_id}"
    )
    assert response.status_code == 422


@pytest.mark.asyncio
async def test_delete_record_previous_publ_403(
    authenticated_client: AsyncClient,
    session_maker: Callable[[], AsyncSession],
    seed_data: SeedData,
) -> None:
    """Test that deleting a record from a previous publication returns 403."""
    user = seed_data["users"][0]
    async with session_maker() as session:
        now = datetime.now(UTC).replace(tzinfo=None)
        record = EventRecord(
            id=uuid4(),
            user_id=user.user_id,
            publ_id=seed_data["publs"][1].id,
            type="rec_ok",
            created_at=now,
            updated_at=now,
        )
        session.add(record)
        await session.commit()

        response = await authenticated_client.delete(
            f"/api/records/{record.id}?user_id={user.user_id}"
        )
        assert response.status_code == status.HTTP_403_FORBIDDEN
        assert response.json()["message"] == "Cannot modify record"


@pytest.mark.asyncio
async def test_update_record_wrong_user(
    authenticated_client_user2: AsyncClient, seed_data
):
    """Test that updating another user's record returns 403."""
    user2 = seed_data["users"][1]
    record_id = seed_data["record_ids"][0]
    response = await authenticated_client_user2.put(
        f"/api/records/{record_id}?user_id={user2.user_id}",
        json={"publ_id": seed_data["publs"][0].id, "latitude": 55.5},
    )
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_export_records_default(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test export returns Excel file with user's records."""
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records/export?user_id={user.user_id}"
    )
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_export_records_project_non_admin_403(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test export with scope=project returns 403 for non-admin."""
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records/export?user_id={user.user_id}&scope=project"
    )
    assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.asyncio
async def test_export_records_csv(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test export returns CSV file with user's records."""
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records/export?user_id={user.user_id}&format=csv"
    )
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    assert "attachment" in response.headers["content-disposition"]
    assert "records.csv" in response.headers["content-disposition"]
    content = response.text
    assert "Genus" in content or "Family" in content


@pytest.mark.asyncio
async def test_export_records_xlsx_default(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test export returns Excel file by default."""
    user = seed_data["users"][0]
    response = await authenticated_client.get(
        f"/api/records/export?user_id={user.user_id}"
    )
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "records.xlsx" in response.headers["content-disposition"]


def _create_test_excel_content() -> bytes:
    """Create a minimal Excel file with test record data."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    from service.export import COLUMN_MAPPING

    headers = list(COLUMN_MAPPING.values())
    ws.append(headers)

    ws.append(["Family", "Genus", "Species", "Country", "55.5", "37.5"])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _create_test_csv_content() -> bytes:
    """Create a minimal CSV file with test record data."""
    from service.export import COLUMN_MAPPING

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(list(COLUMN_MAPPING.values()))
    writer.writerow(["Family", "Genus", "Species", "Country", "55.5", "37.5"])
    return output.getvalue().encode("utf-8")


@pytest.mark.asyncio
async def test_import_from_excel(
    authenticated_client: AsyncClient, seed_data, session_maker
) -> None:
    """Test importing records from Excel file."""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    from service.export import COLUMN_MAPPING

    headers = list(COLUMN_MAPPING.values())
    ws.append(headers)

    row_data = [None] * len(headers)
    header_to_idx = {h: i for i, h in enumerate(headers)}
    if "Family" in header_to_idx:
        row_data[header_to_idx["Family"]] = "TestFamily"
    if "Genus" in header_to_idx:
        row_data[header_to_idx["Genus"]] = "TestGenus"
    if "Species" in header_to_idx:
        row_data[header_to_idx["Species"]] = "TestSpecies"
    if "Country" in header_to_idx:
        row_data[header_to_idx["Country"]] = "TestCountry"
    if "Latitude" in header_to_idx:
        row_data[header_to_idx["Latitude"]] = 55.5
    if "Longitude" in header_to_idx:
        row_data[header_to_idx["Longitude"]] = 37.5

    ws.append(row_data)

    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()

    response = await authenticated_client.post(
        "/api/records/import",
        files={"file": ("test.xlsx", excel_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert "imported" in data
    assert "failed" in data
    assert data["imported"] >= 1


@pytest.mark.asyncio
async def test_import_from_csv(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test importing records from CSV file."""
    from service.export import COLUMN_MAPPING

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(list(COLUMN_MAPPING.values()))

    headers = list(COLUMN_MAPPING.values())
    row_data = [None] * len(headers)
    header_to_idx = {h: i for i, h in enumerate(headers)}
    if "Family" in header_to_idx:
        row_data[header_to_idx["Family"]] = "TestFamily"
    if "Genus" in header_to_idx:
        row_data[header_to_idx["Genus"]] = "TestGenus"
    if "Species" in header_to_idx:
        row_data[header_to_idx["Species"]] = "TestSpecies"
    if "Country" in header_to_idx:
        row_data[header_to_idx["Country"]] = "TestCountry"
    if "Latitude" in header_to_idx:
        row_data[header_to_idx["Latitude"]] = "55.5"
    if "Longitude" in header_to_idx:
        row_data[header_to_idx["Longitude"]] = "37.5"

    writer.writerow(row_data)
    csv_content = output.getvalue().encode("utf-8")

    response = await authenticated_client.post(
        "/api/records/import",
        files={"file": ("test.csv", csv_content, "text/csv")},
    )
    assert response.status_code == 200
    data = response.json()
    assert "imported" in data
    assert "failed" in data
    assert data["imported"] >= 1


@pytest.mark.asyncio
async def test_import_invalid_file_type(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test importing invalid file type returns error."""
    response = await authenticated_client.post(
        "/api/records/import",
        files={"file": ("test.txt", b"invalid content", "text/plain")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["imported"] == 0


@pytest.mark.asyncio
async def test_import_empty_file(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test importing empty Excel file."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    from service.export import COLUMN_MAPPING

    ws.append(list(COLUMN_MAPPING.values()))

    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()

    response = await authenticated_client.post(
        "/api/records/import",
        files={"file": ("empty.xlsx", excel_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["imported"] == 0
    assert data["failed"] == 0


@pytest.mark.asyncio
async def test_import_boolean_fields(
    authenticated_client: AsyncClient, seed_data
) -> None:
    """Test importing records with boolean fields."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    from service.export import COLUMN_MAPPING

    headers = list(COLUMN_MAPPING.values())
    ws.append(headers)

    row_data = [None] * len(headers)
    header_to_idx = {h: i for i, h in enumerate(headers)}
    if "Family" in header_to_idx:
        row_data[header_to_idx["Family"]] = "TestFamily"
    if "Genus" in header_to_idx:
        row_data[header_to_idx["Genus"]] = "TestGenus"
    if "Species" in header_to_idx:
        row_data[header_to_idx["Species"]] = "TestSpecies"
    if "Manual Location" in header_to_idx:
        row_data[header_to_idx["Manual Location"]] = "TRUE"
    if "Date Interval" in header_to_idx:
        row_data[header_to_idx["Date Interval"]] = "YES"
    if "Taxon Verbatim" in header_to_idx:
        row_data[header_to_idx["Taxon Verbatim"]] = "FALSE"

    ws.append(row_data)

    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()

    response = await authenticated_client.post(
        "/api/records/import",
        files={"file": ("test.xlsx", excel_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["imported"] >= 1


@pytest.mark.asyncio
async def test_import_user_without_publ(
    authenticated_client_user2: AsyncClient, seed_data
) -> None:
    """Test importing with user that has no publication returns error."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    from service.export import COLUMN_MAPPING

    ws.append(list(COLUMN_MAPPING.values()))

    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()

    response = await authenticated_client_user2.post(
        "/api/records/import",
        files={"file": ("test.xlsx", excel_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["imported"] == 0


@pytest.mark.asyncio
async def test_import_file_size_limit(
    authenticated_client: AsyncClient, seed_data, monkeypatch
) -> None:
    """Test importing file that exceeds size limit."""
    from core.config import settings

    large_content = b"x" * (settings.MAX_IMPORT_FILE_BYTES + 1)

    response = await authenticated_client.post(
        "/api/records/import",
        files={"file": ("large.xlsx", large_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["imported"] == 0


@pytest.mark.asyncio
async def test_import_limit_enforcement(
    authenticated_client: AsyncClient, seed_data, session_maker, monkeypatch
) -> None:
    """Test that import limit is enforced."""
    from core.config import settings

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    from service.export import COLUMN_MAPPING

    headers = list(COLUMN_MAPPING.values())
    ws.append(headers)

    header_to_idx = {h: i for i, h in enumerate(headers)}

    num_records = settings.MAX_RECORDS_PER_PUBLICATION + 1
    for _ in range(num_records):
        row_data = [None] * len(headers)
        if "Family" in header_to_idx:
            row_data[header_to_idx["Family"]] = "TestFamily"
        if "Genus" in header_to_idx:
            row_data[header_to_idx["Genus"]] = "TestGenus"
        if "Species" in header_to_idx:
            row_data[header_to_idx["Species"]] = "TestSpecies"
        ws.append(row_data)

    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()

    response = await authenticated_client.post(
        "/api/records/import",
        files={"file": ("test.xlsx", excel_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["imported"] == 0
